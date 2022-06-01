import requests
import openpyxl
import configparser


config=configparser.RawConfigParser()
config.read('config.ini',encoding='utf-8-sig')
obs_cookie=config.get('default','obs_cookie')
config=configparser.ConfigParser(interpolation=configparser.BasicInterpolation())
config.read('config.ini',encoding='utf-8-sig')
recordFileName=config.get('default','record_file')
sheetName=config.get('default','sheetname')
domain=config.get('default','obs_domain')
project=config.get('default','project')
project=project.replace(":","%3A")


wb = openpyxl.load_workbook(recordFileName)
sheet=wb[sheetName]
column=sheet.max_column+1
sheet.cell(1,column).value='needCommitId'

for row in range(2,sheet.max_row+1):
    package=sheet.cell(row,1).value

    url=domain+'/package?project='+project+'&draw=6' \
    '&columns%5B0%5D%5Bdata%5D=name&columns%5B0%5D%5Bname%5D=&columns%5B0%5D%5Bsearchable%5D=true&columns%5B0%5D%5Borderable%5D=true' \
        '&columns%5B0%5D%5Bsearch%5D%5Bvalue%5D=&columns%5B0%5D%5Bsearch%5D%5Bregex%5D=false&columns%5B1%5D%5Bdata%5D=changed&columns%5B1%5D%5Bname%5D=' \
            '&columns%5B1%5D%5Bsearchable%5D=true&columns%5B1%5D%5Borderable%5D=true&columns%5B1%5D%5Bsearch%5D%5Bvalue%5D=&columns%5B1%5D%5Bsearch%5D%5Bregex%5D=false' \
                '&order%5B0%5D%5Bcolumn%5D=0&order%5B0%5D%5Bdir%5D=asc&start=0&length=25&search%5Bvalue%5D='+package+'&search%5Bregex%5D=false&_=1651115738457'

    
    headers = {
        "cookie": obs_cookie
        }
    response=requests.get(url,headers=headers).json()
    if response['recordsFiltered']==0:
        sheet.cell(row,column).value='1'

    

wb.save(recordFileName)