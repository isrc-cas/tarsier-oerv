import openpyxl
import requests
import re
import configparser


config=configparser.ConfigParser(interpolation=configparser.BasicInterpolation())
config.read('config.ini',encoding='utf-8-sig')
recordFileName=config.get('default','record_file')
sheetName=config.get('default','sheetname')
gitee_token=config.get('default','gitee_token')


wb = openpyxl.load_workbook(recordFileName)
sheet=wb[sheetName]
column=sheet.max_column+1
sheet.cell(1,column).value='commit'
for i in range(1,column):
    #print (sheet.cell(1,i).value)
    if sheet.cell(1,i).value=='needCommitId':
        needFlag=i
        break
#print(sheet.max_row+1)
for row in range(2,sheet.max_row+1):
    if 'needFlag' not in dir() or sheet.cell(row,needFlag).value=='1':
        package=sheet.cell(row,1).value
        print ('get commit info for:'+package)
        branch=config.get('default','branch')
        url='https://gitee.com/api/v5/repos/src-openeuler/'+package+'/branches/'+branch+'?access_token='+gitee_token
        r = requests.get(url)
        #print (r.text)
        if re.search('Branch does not exist',r.text):
            url='https://gitee.com/api/v5/repos/src-openeuler/'+package+'/branches/master?access_token='+gitee_token
            r = requests.get(url)
            branch='master'
        latestCommit=r.json()['commit']['sha']
        #check whether has spec
        url='https://gitee.com/api/v5/repos/src-openeuler/'+package+'/contents/%2F?access_token='+gitee_token+'&ref='+branch
        print(url)
        r = requests.get(url).text
        print (r)
        print ('')
        if '.spec' in r:
            sheet.cell(row,column).value=latestCommit

wb.save(recordFileName)