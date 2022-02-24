import requests
import json
import re
import time
import openpyxl
import configparser

my_cookie=''

def searchPackage(project,keyword,start=0,length=25):
    project=project.replace(":","%3A")
    #print('project:'+project)
    now=str(int(round(time.time() * 1000)))
    searchRequest= 'https://build.openeuler.org/package?project='+project+'&draw=1&columns%5B0%5D%5Bdata%5D=name&columns%5B0%5D%5Bname%5D=&columns%5B0%5D%5Bsearchable%5D=true&columns%5B0%5D%5Borderable%5D=true&columns%5B0%5D%5B' \
            'search%5D%5Bvalue%5D=&columns%5B0%5D%5Bsearch%5D%5Bregex%5D=false&columns%5B1%5D%5Bdata%5D=changed&columns%5B1%5D%5Bname%5D=&columns%5B1%5D%5Bsearchable%5D=true' \
                '&columns%5B1%5D%5Borderable%5D=true&columns%5B1%5D%5Bsearch%5D%5Bvalue%5D=&columns%5B1%5D%5Bsearch%5D%5Bregex%5D=false&order%5B0%5D%5Bcolumn%5D=0&order%5B0%5D%5Bdir%5D=asc' \
                    '&start='+str(start)+'&length='+str(length)+'&search%5Bvalue%5D='+keyword+'&search%5Bregex%5D=false&_='+now
    #print(searchRequest)
    response=requests.get(searchRequest)
    data=json.loads(response.text)
    #print(data)
    return data

def recordPackage(ws,project,keyword):
    data=searchPackage(project,keyword)
    recordsFiltered=data['recordsFiltered']
    print("find packages related to "+keyword+":"+str(recordsFiltered))

    length=25
    if recordsFiltered%length==0:
        totalPage=recordsFiltered//length
    else:
        totalPage=recordsFiltered//length+1
    
    num=0
    ws.append(['PackageName'])
    for page in range(totalPage):
        start=page*length
        data=searchPackage(project,keyword,start,length)

        packageList = []
        
        for package in data['data']:
            name=re.search('>(.*)<',package['name']).group(1)
            #print (name)
            packageList.append(name)
        print (packageList)

        for package in packageList:
            ws.append([package])
            num+=1
    
    if num!=recordsFiltered:
        print("there should be "+str(recordsFiltered)+" packages")
        print("actually find "+str(num)+" packages")
        print("please check")
        raise Exception ("")

def hasRPM(project,package,repo):
    #print(my_cookie)
    requestURL='https://build.openeuler.org/package/binaries/'+project+'/'+package+'/'+repo
    headers = {
        "cookie": my_cookie
        }
    response=requests.get(requestURL,headers=headers)
    #print(response.text)
    if re.search('No built binaries',response.text):
        print (package+' never build successfully')
        return False
    elif re.search(package+'[^<>]+\.rpm"><i class=\'fas fa-download text-secondary\'>',response.text):
        print (package+' has rpm')
        return True
    else:
        print ("can't check whether has rpm")
        raise Exception ("")

def recordHasRPM(ws,beginrow,project,repo):
    column=ws.max_column+1
    ws.cell(beginrow-1,column).value=project+'_HasRPM'
    for row in range(beginrow,ws.max_row+1):
        package=ws.cell(row,1).value
        try:
            if hasRPM(project,package,repo):
                ws.cell(row,column).value='Yes'
            else:
                ws.cell(row,column).value='No'
        except:
            continue




# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    now=time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()) 
    recordFileName='PackageHasRPM_'+now+'.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active

    config=configparser.RawConfigParser()
    config.read('config.ini',encoding='utf-8-sig')
    project=config.get('default','project')
    print(project)
    repo=config.get('default','repo')
    print(repo)
    target_package=config.get('default','target_package')
    print(target_package)
    my_cookie=config.get('default','cookie')
    
    recordPackage(ws,project,target_package)
    recordHasRPM(ws,2,project,repo)
    wb.save(recordFileName)


