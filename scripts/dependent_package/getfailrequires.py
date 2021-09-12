import json
import os
from openpyxl import Workbook
from openpyxl import load_workbook

packaglistfilepath = os.path.join(os.getcwd(), 'packagelist.json')
buildrequire_filepath = os.path.join(os.getcwd(), 'buildrequiresfile.json')
excelfile = os.path.join(os.getcwd(), 'packagelist.xlsx')

def get_requires(packagelist, datafile):
    with open(datafile, 'r') as f:
        requiresdata = json.load(f)

    requireslist = []
    for pkg in packagelist:
        for item in requiresdata:
            if pkg == item['package']:
               requireslist = requireslist + item['buildrequires']
    requireslist = list(set(requireslist))
    print ('requireslist', len(requireslist))
    return requireslist


def compare_requires(faillist, passlist):
    comparelist = [x for x in faillist if x not in passlist]
    print ('comparelist', len(comparelist))
    return comparelist

def create_excelfile(packagelistfile, requiresdata, comparedata, excelfile):
    with open(packagelistfile, 'r') as f:
        packagelistdata = json.load(f)
    passpackagelist = packagelistdata['succeeded']
    wb = Workbook()
    ws = wb.active
    ws.title = 'succeeded packages'
    wb.create_sheet("unresolvable packages requires",1)
    wb.create_sheet("unresolvable requires not in succeeded",2)
    wb.save(excelfile)
    wb = load_workbook(excelfile)
    ws = wb.worksheets[0]
    ws.column_dimensions["A"].width = 45
    for i in range(len(passpackagelist)):
        ws.cell(i+1,1).value = passpackagelist[i]
    ws = wb.worksheets[1]
    ws.column_dimensions["A"].width = 50
    for j in range(len(requiresdata)):
        ws.cell(j+1,1).value = requiresdata[j]
    ws = wb.worksheets[2]
    ws.column_dimensions["A"].width = 50
    for k in range(len(comparedata)):
        ws.cell(k+1,1).value = requiresdata[k]

    wb.save(excelfile)


if __name__=="__main__":
    with open(packaglistfilepath, 'r') as f:
        packagelistdata = json.load(f)
    failrequireslist = get_requires(packagelistdata['unresolvable'], buildrequire_filepath)
    passrequireslist = get_requires(packagelistdata['succeeded'], buildrequire_filepath)
    compare_list = compare_requires(failrequireslist, passrequireslist)
    create_excelfile(packaglistfilepath, failrequireslist, compare_list, excelfile)
