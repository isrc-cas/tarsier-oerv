#!/usr/bin/env python3
# -*-coding:utf-8-*-

import requests
import json
import re
import os
import constant
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import worksheet
from openpyxl.styles import Side, Border
import time
import datetime

base_url = 'https://build.openeuler.org/project/monitor/openEuler:Mainline:RISC-V?arch_riscv64=1&defaults=0&repo_standard_riscv64=1'

def get_packagelist(url):
    response = requests.get(url)
    data = response.text
    datalist = data.split('\n')
    pacakgedatalist = [x for x in datalist if x.startswith('<tbody')]
    packagedata = pacakgedatalist[0].split("='")
    data_packagenames = packagedata[1].split("]'")[0].replace("[","").replace("&quot;","").split(", ")

    return data_packagenames

# 定义状态对应的编号（用于excel数据分析和sheetindex）
def getCode(status):
    code = 6
    if status == 'succeeded':
        code = 0
    if status == 'failed':
        code = 1
    if status == 'unresolvable':
        code = 2
    if status == 'broken':
        code = 3
    if status == 'disabled':
        code = 4
    if status == 'excluded':
        code = 5
    if status == 'blocked':
        code = 6
    if status == 'scheduled':
        code = 7
    if status == 'building':
        code = 8

    return code

# 如果文件obsBuild不存在则新建一个
def createSheetFileTotal(filepath):
    try:
        wb = load_workbook(obsTotalNumfilepath)
        ws = wb['buildresultsTotal']
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.create_sheet("buildresultsTotal", 0)  #记录统计信息
        ws.append(['datetime',"succeeded","failed","unresolvable","broken","disabled","excluded","blocked","scheduled","building"])
        wb.save(filepath)

# 按照构建时间新建一个表记录构建详细记录和状态
def createSheetFileDetial(filepath, nowtime):
    wb = Workbook()
    ws = wb.create_sheet(nowtime, 0)  # 按日期记录详细构建信息
    ws.append(['packagename', 'statuscode', 'statusname'])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    wb.save(filepath)

def writeDetialToExcle(filepath, datelist, dt, title):
    wb = load_workbook(filepath)
    ws = wb[dt]
    for itPkgname in datelist:
        ws.append([itPkgname, getCode(title), title])

    print(title + ": ",len(datelist))
    wb.save(filepath)

def writeTotalToExcle(filepath, data):
    wb = load_workbook(filepath)
    ws = wb['buildresultsTotal']
    ws.append(data)
    wb.save(filepath)



if __name__=="__main__":
    # 根据日期新建excel，并写入数据。用于归档和数据分析用
    nowtime = time.strftime("%Y%m%d-%H%M", time.localtime())
    print(nowtime)

    obsdatafolder = os.getcwd().replace("scripts\dependent_package",'data\obsBuildStatus')
    obsTotalNumfilepath = os.path.join(obsdatafolder, 'obsBuild.xlsx')
    obsDetialfilepath = os.path.join(obsdatafolder, 'obsBuild-'+nowtime+'.xlsx')

    createSheetFileTotal(obsTotalNumfilepath)
    createSheetFileDetial(obsDetialfilepath, nowtime)

    totalNum = [nowtime]
    # 获取各状态下的packagesname并写入数据到excle
    queryDate = ["succeeded","failed","unresolvable","broken","disabled","excluded","blocked","scheduled","building"]
    for qd in queryDate:
        url = base_url+"&"+qd+"=1"
        # print(qd,"url=",url)
        packagelist = get_packagelist(url)
        writeDetialToExcle(obsDetialfilepath, packagelist, nowtime, qd)
        totalNum.append(len(packagelist))

    writeTotalToExcle(obsTotalNumfilepath, totalNum)

    print('All datas in file ：', obsDetialfilepath)