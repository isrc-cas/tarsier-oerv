import requests
import json
import base64
import re
import os
import constant
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Side, Border
import time
import datetime

token = constant.token
headers = {
    'Content-Type': 'application/json;charset=UTF-8'
    }
params = {
    'access_token': token
}

specfilepath = os.path.join(os.getcwd(), 'specfile.json')

obsdatafolder = os.getcwd().replace("scripts\dependent_package", 'data')
dependentsfilepath = os.path.join(obsdatafolder, 'dep_info.xlsx')
specContent = os.path.join(os.getcwd(), 'debugjson/speccontent.txt')

def createSheet():
    nowtime = time.strftime("%Y%m%d%H%M", time.localtime())
    print(nowtime)

    wb = Workbook()
    wb.create_sheet('pkg' + nowtime, 0)
    wb.create_sheet('dep' + nowtime, 1)
    wb.save(dependentsfilepath)


def  writePkgInfo2Excle(requirepakcage_list):
    wb = load_workbook(dependentsfilepath)
    # ws = wb.worksheets[nowtime]
    ws = wb.worksheets[0]
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 100
    ws.append(['package', 'packagename', 'buildrequires'])

    for item in requirepakcage_list:
        ws.append([item[0],item[1],','.join(item[2])])

    print("writePkgInfo2Excle is done! ")
    wb.save(dependentsfilepath)


def writeDependentInfo2Excle(requirepakcage_list):
    wb = load_workbook(dependentsfilepath)
    ws = wb.worksheets[1]
    # ws = wb[nowtime]
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.append(['package', 'packagename', 'buildrequires'])

    for item in requirepakcage_list:
        for pkg in item[2]:
            row = [item[0], item[1], pkg]
            ws.append(row)

    print("writeDependentInfo2Excle is done!  file:"+dependentsfilepath)
    wb.save(dependentsfilepath)


def get_buildrequires():
    with open(specfilepath, 'r') as f:
        specdata = json.load(f)

    requirepakcage_list = []
    for specfile in specdata:
        # if语句用于调试的时候，输出一个数据看数据信息用
        # if specdata.index(specfile) > 0:
        #     break

        # 获取spec文件的内容，并进行转码
        response = requests.get(specfile['apiurl'], headers=headers, params=params)
        specfiledata = json.loads(response.text)
        spec_content = base64.b64decode(specfiledata['content']).decode()

        # 调试用：将转码后的content写入debugjson/speccontent.json；便于分析数据结构
        # with open(specContent, 'w') as f:
        #     f.write(spec_content)
        #     f.close()

        # 对spec文件进行解读，分析BuildRequires
        specdata_list = spec_content.split('\n')
        namelist = [x for x in specdata_list if re.match('[Nn]ame *:', x)]
        namelist = [re.sub(r'\t', ' ', x) for x in namelist]
        namelist = [re.sub(r' ', '', x) for x in namelist]
        # print ('namelist', namelist)

        packagename = list(filter(None, namelist[0].split(':')))[1]
        # print ('packagename', packagename)

        buildrequires = [x for x in specdata_list if x.startswith('BuildRequires:')]
        buildrequires = [x.replace('BuildRequires:', 'BuildRequires: ') for x in buildrequires]
        buildrequires = [re.sub(r'\t', ' ', x) for x in buildrequires]
        # print ('buildrequire', buildrequires)

        requirespkg_list = []
        for item in buildrequires:
            itemlist = list(filter(None, item.split(' ')))
            for i in range(len(itemlist)):
                m1 = re.match(r'^[^A-Za-z0-9_/-]*', itemlist[i]).group()
                m2 = re.match(r'^(\d\W)*', itemlist[i]).group()
                if m1 or m2:
                    itemlist[i] = ''
            itemlist.remove('BuildRequires:')
            itemlist = list(filter(None, itemlist))
            requirespkg_list = requirespkg_list + itemlist

        requirepakcage_list.append([specfile['package'], packagename, requirespkg_list])

    return requirepakcage_list


if __name__=="__main__":
    requirepakcage_list = get_buildrequires()
    createSheet()
    writePkgInfo2Excle(requirepakcage_list)
    writeDependentInfo2Excle(requirepakcage_list)



