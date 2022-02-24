import requests
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import json
import constant
from openpyxl.styles import Side, Border, PatternFill, Font, Alignment
import datetime


def get_giteepkg(headers,token,owner):
    i = 1
    pkglist = []
    while True:
        pkg_params = {
            'access_token': token,
            'state': 'all',
            'page': i,
            'per_page': 100
        }
        pkg_url = 'https://gitee.com/api/v5/orgs/{}/repos'.format(owner)
        pkg_resp = requests.get(pkg_url, headers=headers, params=pkg_params)
        pkg_data = json.loads(pkg_resp.text)
        # print ('pkg_data', pkg_data)
        if len(pkg_data) > 0:
            print ('page', i)
            for pkg in pkg_data:
                pkglist.append(pkg['name'])
            i = i + 1
        else:
            break
    # print (pkglist)
    print ('pkglist length', len(pkglist))
    return pkglist

def get_PRInfo(pkgs,headers,token,owner):
    prinfo_list = []
    for pkg in pkgs:
        print('pkg',pkg)
        prlist = []
        i = 1
        while True:
            pr_params = {
                'access_token': token,
                'state': 'all',
                'sort': 'created',
                'direction': 'desc',
                'page': i,
                'per_page': 100
            }
            pr_url = 'https://gitee.com/api/v5/repos/{}/{}/pulls'.format(owner,pkg)
            pr_resp = requests.get(pr_url, headers=headers, params=pr_params)
            pr_data = json.loads(pr_resp.text)
            # print ('pr_data', pr_data)
            if len(pr_data) > 0:
               prlist.extend(pr_data)
               i = i + 1
            else:
               break
        # print ('prlist length', len(prlist))
        if len(prlist) > 0:
            created_time = datetime.datetime.strptime(prlist[0]['created_at'],"%Y-%m-%dT%H:%M:%S+08:00").strftime("%Y-%m-%d %H:%M:%S")
            updated_time = datetime.datetime.strptime(prlist[0]['updated_at'],"%Y-%m-%dT%H:%M:%S+08:00").strftime("%Y-%m-%d %H:%M:%S")
            j = 1
            comments_list = []
            while True:
                comments_params = {
                'access_token': token,
                'page': j,
                'per_page': 20,
                'direction': 'desc'
                }
                comments_url = 'https://gitee.com/api/v5/repos/{}/{}/pulls/{}/comments'.format(owner,pkg,len(prlist))
                comments_resp = requests.get(comments_url, headers=headers, params=comments_params)
                comments_data = json.loads(comments_resp.text)
                # print ('comments_data', comments_data)
                if len(comments_data) > 0:
                    comments_list.extend(comments_data)
                    j = j + 1
                else:
                    break
            for item in comments_list:
                if item['user']['name'] != 'openeuler-ci-bot':
                    lastest_comment_time = datetime.datetime.strptime(item['updated_at'],"%Y-%m-%dT%H:%M:%S+08:00").strftime("%Y-%m-%d %H:%M:%S")
                    lastest_comment_user = item['user']['name']
                    break
                else:
                    lastest_comment_time = ''
                    lastest_comment_user = ''
                # print('lastest_comment', lastest_comment_time,lastest_comment_user)
            prinfo = [
                pkg, 
                prlist[0]['user']['name'],
                prlist[0]['html_url'],
                prlist[0]['state'],
                created_time,
                updated_time,
                lastest_comment_time,
                lastest_comment_user,
                ]
        else:
            prinfo = [pkg,'','','','','','','']
        prinfo_list.append(prinfo)
    print('prinfo_list length', len(prinfo_list))
    return prinfo_list

def create_excelfile(prinfo_list,report_header,excelfile):
    wb = Workbook()
    ws = wb.active
    ws.title = "openEuler-RISC-V PR Info"
    wb.save(excelfile)
    wb = load_workbook(excelfile)
    ws = wb.worksheets[0]
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 30
    titlefont = Font(name='Arial', size=11, bold=True, italic=False, color='000000')
    titlefill = PatternFill(fill_type="solid",fgColor="CCCCFF")
    urlfont = Font(name='Arial', size=11, bold=False, italic=False, underline='single', color='3366FF')
    ws.append(report_header)
    for i,j in enumerate(prinfo_list):
        for m,n in enumerate(j):
            if m != 2:
                ws.cell(column=m+1,row=i+2,value=n)
            else:
                ws.cell(column=m+1,row=i+2).hyperlink = n
    for k in range(1,len(report_header)+1):
        ws.cell(row=1,column=k).fill = titlefill
        ws.cell(row=1,column=k).font = titlefont
        ws.cell(row=1,column=k).alignment = Alignment(horizontal='center', vertical='center')
    for r in range(2,len(prinfo_list)+2):
        ws.cell(row=r,column=3).font = urlfont
    max_row = ws.max_row
    max_column = ws.max_column
    side = Side(border_style='thin', color='000000')
    border = Border(left=side,right=side,top=side,bottom=side)
    for p in range(1, max_row+1):
        for q in range(1, max_column+1):
            ws.cell(p, q).border = border
                
    wb.save(excelfile)



    



if __name__=="__main__":
    token = constant.token
    headers = constant.headers
    owner = constant.owner
    pkg_list = get_giteepkg(headers,token,owner)
    prinfo_list = get_PRInfo(pkg_list,headers,token,owner)
    report_header = constant.report_header
    excelfile = constant.excelfile
    create_excelfile(prinfo_list,report_header,excelfile)

    