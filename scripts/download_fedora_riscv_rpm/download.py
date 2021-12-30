import os
import wget
import para
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Side, Border
import json
import shutil

excelfilepath = para.excelfile
report_header = para.report_header
rpminfo_filepath = para.rpminfo_filepath


def rpm_download(rpminfo_file):
    with open(rpminfo_file, 'r') as f:
        pkgdata_list = json.load(f)
    
    reportdata_list = []
    for pkgdata in pkgdata_list:
        if pkgdata['result'] == 'success':
           filedir = os.path.join(os.getcwd(), 'downloaded_rpm', pkgdata['package'])
           if os.path.exists(filedir):
               shutil.rmtree(filedir)
           os.makedirs(filedir)
           reportdata = []
           print ('{} downloading.'.format(pkgdata['package']))
           for rpmdata in pkgdata['rpmlist']:
               filename = rpmdata['rpm_name']
               filepath = os.path.join(filedir, filename)
               try:
                   wget.download(rpmdata['rpm_url'], filepath)
                   result = 'success'
               except:
                   result = 'fail'
               reportdata = [pkgdata['package'],rpmdata['rpm_name'],result]
               reportdata_list.append(reportdata)
        else:
            reportdata = [pkgdata['package'],'none','fail']
            reportdata_list.append(reportdata)

    testfile = os.path.join(os.getcwd(), 'data.json')
    
    return reportdata_list



def create_excelfile(reportdata, header, excelfile):
    wb = Workbook()
    ws = wb.active

    ws.title = "pkg_rpm"
    wb.save(excelfile)
    wb = load_workbook(excelfile)
    ws = wb.worksheets[0]
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 20
    ws.append(header)
    for row in reportdata:
        ws.append(row)
    max_row = ws.max_row
    max_column = ws.max_column
    side = Side(border_style='thin', color='FF000000')
    border = Border(left=side,right=side,top=side,bottom=side)
    for i in range(1, max_row+1):
        for j in range(1, max_column+1):
            ws.cell(i, j).border = border
    
    wb.save(excelfile)


if __name__=="__main__":
    report_data = rpm_download(rpminfo_filepath)
    create_excelfile(report_data, report_header, excelfilepath)