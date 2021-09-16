
import requests
import re
import configparser
import datetime
import os
from openpyxl import Workbook

'''
def parseWatchList():
    pullIdList=[]
    with open(getConfig("default", "watchList"),mode='r') as F:
        list=F.readlines()
        for url in list:
            try:
                pullId=re.search("pulls/(\\d+)",url).group(1)
                pullIdList.append(pullId)
            except:
                print("can't find pull id in "+url)
                continue

    return pullIdList
'''
curDir=os.path.dirname(os.path.abspath(__file__))


def getConfig(section,option):
    configFile=curDir+"/config.ini"
    config = configparser.ConfigParser()
    config.read(configFile, encoding="utf-8")
    return config.get(section, option)

def getRepoUrlInOrg(org,token):
    i=1
    urlList=[]
    while 1:
        url='https://gitee.com/api/v5/orgs/'+org+'/repos?access_token='+token+'&type=all&page='+str(i)+'&per_page=100'
        print("Get repos under organization "+org+"...")
        print (url)
        r = requests.get(url).json()
        for e in r:
            urlList.append(e['url'])
        if len(r)<100:
            break
        else:
            i+=1
    return urlList

def getOwnerAndRepo(urlList):
    ownerAndRepo=[]
    for url in urlList:
        result=re.search('/([^/]*)/([^/]*)$',url)
        ownerAndRepo.append([result.group(1),result.group(2)])
    #print(ownerAndRepo)
    return ownerAndRepo


def getPRListInRepo(owner,repo,token):
    i=0
    PRList=[]
    while 1:
        url="https://gitee.com/api/v5/repos/"+owner+"/"+repo+"/pulls?access_token="+token+"&state=all&sort=created&direction=asc&page=1&per_page=100"
        print("Get "+url+"...")
        r=requests.get(url).json()
        if len(r)>0:
            for e in r:
                #print(e['number'])
                PRList.append(e['number'])
            print(PRList)
        else:
            print('there is no PR under '+owner+'/'+repo)
        if len(r)<100:
            break
        else:
            i+=1
    return PRList


def getPRInfo(owner,repo,pullId,token):
    PRInfo={}

    #get PR info
    url = "https://gitee.com/api/v5/repos/"+owner+"/"+repo+"/pulls/"+str(pullId)+"?access_token="+token
    print("Get "+url+"...")

    r=requests.get(url).json()
    #print(r)
    PRInfo['title']=r['title']
    PRInfo['state']=r['state']
    PRInfo['url']=r['html_url']

    #get PR comment
    i=0
    num=0
    ignoreComment=getConfig("default","ignoreComment")
    ignoreComment=ignoreComment.split(",")
    while 1:
        url="https://gitee.com/api/v5/repos/"+owner+"/"+repo+"/pulls/"+str(pullId)+"/comments?access_token="+token+"&page="+str(i)+"&per_page=100&direction=desc"
        print("Get " + url + "...")
        r = requests.get(url).json()

        #print(r)
        # if it is a robot reply,continue,else record the last comment by human
        #if comment in one day,num+1
        for e in r:

            if e['user']['name'] in ignoreComment:
                #print(e['user']['login'])
                continue
            else:
                if 'lastCommentHuman' not in PRInfo.keys():
                    PRInfo['lastCommentHuman']=e['user']['name']
                    PRInfo['lastComment']=e['body'].replace("\n", "")
                updateTime=e['updated_at']
                updateTime=re.search('(.*)\\+08:00',updateTime).group(1)
                updateTime=updateTime.replace('T',' ')
                updateTime=datetime.datetime.strptime(updateTime, '%Y-%m-%d %H:%M:%S')
                #print(updateTime)
                now=datetime.datetime.now()
                if (now-updateTime).days<1:
                    num+=1
        #if the num of comments <100, is last page
        if len(r)<100:
            break
    if 'lastCommentHuman' not in PRInfo.keys():
        PRInfo['lastCommentHuman']=""
    if 'lastComment' not in PRInfo.keys():
        PRInfo['lastComment']=""
    PRInfo['commentsInOneDay']=num

    return PRInfo



def getPRInfoByUrl(urlList):
    ownerAndRepo = getOwnerAndRepo(urlList)
    allPRInfo=[]
    for e in ownerAndRepo:
        PRList = getPRListInRepo(e[0], e[1], token)
        for PR in PRList:
            PRInfo = getPRInfo(e[0], e[1], PR, token)
            #print (PRInfo)
            allPRInfo.append(PRInfo)
    #print(allPRInfo)
    return allPRInfo




def writeExcel(ws,PRInfo):
    for pr in PRInfo:
        ws.append([pr['title'],pr['url'],pr['state'],pr['lastCommentHuman'],pr['lastComment'],pr['commentsInOneDay']])

if __name__=="__main__":
    #getPRInfoInWatchList()

    token = getConfig("default", "token")
    org=getConfig("default","orgs")
    repo=getConfig("default","repo")

    #create excel
    now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    wb = Workbook()

    #create sheet to store org pr info
    ws = wb.create_sheet(org, 0)
    ws.append(["PRTitle", "PRUrl", "PRState","lastCommentHuman","lastComment","commentsInOneDay"])
    urlList=getRepoUrlInOrg(org,token)
    PRInfo=getPRInfoByUrl(urlList)
    writeExcel(ws,PRInfo)

    # create sheet to store repo pr info
    ws2 = wb.create_sheet(re.search("([^/]*)$",repo).group(1), 0)
    ws2.append(["PRTitle", "PRUrl", "PRState","lastCommentHuman","lastComment","commentsInOneDay"])
    urlList=[repo]
    PRInfo = getPRInfoByUrl(urlList)
    writeExcel(ws2, PRInfo)

    resultFile = curDir + "/PRList_" + now + ".xlsx"
    wb.save(resultFile)



