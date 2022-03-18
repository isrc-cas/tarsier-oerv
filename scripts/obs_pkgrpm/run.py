import requests
import os
import para
from requests.auth import HTTPBasicAuth
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Side, Border

def get_obsdata(account,repo,project):
    pkgs_url = 'https://build.openeuler.org/source/{}'.format(project)
    pkgs_resp = requests.get(pkgs_url,auth=HTTPBasicAuth(account['user'],account['password']))
    pkgs_data = pkgs_resp.text
    pkgslist = re.findall('".*"', pkgs_data)
    del pkgslist[0]
    pkgslist = [x.replace('"','') for x in pkgslist]
    print ('pkgs length', len(pkgslist))
    pkgrpm_list = []
    for pkg in pkgslist:
        print ('obs package', pkg, pkgslist.index(pkg)+1)
        rpmlist = []
        for repo in repolist:
            rpm_url = 'https://build.openeuler.org/build/{}/{}/riscv64/{}'.format(project,repo,pkg)
            rpm_resp = requests.get(rpm_url,auth=HTTPBasicAuth(account['user'],account['password']))
            rpm_data = rpm_resp.text
            # print ('rpm_data', rpm_data)
            rpm_pattern = 'filename=.*.src.rpm'
            try:
                rpm_ver = re.search(rpm_pattern, rpm_data).group()[10:]
                # print ('rpm_ver', rpm_ver)
                rpm_history = 'Yes'
            except:
                rpm_history = 'No'
                rpm_ver = 'None'
                # print ('rpm_ver', rpm_ver)
            rpmlist.extend([rpm_history, rpm_ver])
        pkglist = [pkg] + rpmlist
        pkgrpm_list.append(pkglist)
    return pkgrpm_list


def create_excelfile(reportdata,header,excelfile):
    wb = Workbook()
    ws = wb.active
    ws.title = "OBS Packages rpm Info"
    wb.save(excelfile)
    wb = load_workbook(excelfile)
    ws = wb.worksheets[0]
    ws.column_dimensions["A"].width = 40
    for i in range(66, 65+len(reportdata[0])):
        ws.column_dimensions[chr(i)].width = 60
    ws.append(header)
    for row in reportdata:
        ws.append(row)
    max_row = ws.max_row
    max_column = ws.max_column
    side = Side(border_style='thin', color='FF000000')
    border = Border(left=side,right=side,top=side,bottom=side)
    for i in range(1, max_row+1):
        for j in range(1, max_column+1):
            ws.cell(i, j).border = border
    
    wb.save(excelfile)


if __name__=="__main__":
    obs_account = para.obs_account
    obs_project = para.obs_project
    repolist = para.repolist
    report_data = get_obsdata(obs_account,repolist,obs_project)
    excelheader = para.excelheader
    excelfile = para.excelfile
    create_excelfile(report_data, excelheader, excelfile)   