import requests
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import json
import para
from openpyxl.styles import Side, Border, PatternFill, Font, Alignment
import datetime


def get_giteepkg(headers,token,owner):
    i = 1
    pkglist = []
    while True:
        pkg_params = {
            'access_token': token,
            'state': 'all',
            'page': i,
            'per_page': 100
        }
        pkg_url = 'https://gitee.com/api/v5/orgs/{}/repos'.format(owner)
        pkg_resp = requests.get(pkg_url, headers=headers, params=pkg_params)
        pkg_data = json.loads(pkg_resp.text)
        # print ('pkg_data', pkg_data)
        if len(pkg_data) > 0:
            print ('page', i)
            for pkg in pkg_data:
                pkglist.append(pkg['name'])
            i = i + 1
        else:
            break
    # print (pkglist)
    print ('pkglist length', len(pkglist))
    return pkglist


def get_PRInfo(pkgs,headers,token,owner,period):
    prinfo_list = []
    startdate = datetime.datetime.strptime(period[0],'%Y-%m-%d')
    enddate = datetime.datetime.strptime(period[1],'%Y-%m-%d')
    for pkg in pkgs:
        print('pkg',pkg)
        prlist = []
        i = 1
        while True:
            pr_params = {
                'access_token': token,
                'state': 'all',
                'sort': 'created',
                'direction': 'desc',
                'page': i,
                'per_page': 100
            }
            pr_url = 'https://gitee.com/api/v5/repos/{}/{}/pulls'.format(owner,pkg)
            pr_resp = requests.get(pr_url, headers=headers, params=pr_params)
            pr_data = json.loads(pr_resp.text)
            # print ('pr_data', pr_data)
            if len(pr_data) > 0:
               prlist.extend(pr_data)
               i = i + 1
            else:
               break
        # print ('prlist length', len(prlist))
        if len(prlist) > 0:
            prlist_new = [pr for pr in prlist if startdate <= datetime.datetime.strptime(pr['created_at'],"%Y-%m-%dT%H:%M:%S+08:00") <= enddate]
            for pr in prlist_new:
                created_time = datetime.datetime.strptime(pr['created_at'],"%Y-%m-%dT%H:%M:%S+08:00").strftime("%Y-%m-%d %H:%M:%S")
                prinfo = [pkg, pr['state'], pr['html_url'], pr['user']['name'], created_time]
                prinfo_list.append(prinfo)
        
    print('prinfo_list length', len(prinfo_list))
    return prinfo_list

def create_excelfile(prdata,report_header,excelfile,period):
    period = [x.replace('-','') for x in period]
    userlist = [x[3] for x in prdata]
    userset = set(userlist)
    # print ('userset', userset)
    stat_list = []
    for item in userset:
        stat_list.append([item,userlist.count(item)])
    # print ('stat_list', stat_list)
    wb = Workbook()
    ws = wb.active
    ws.title = "{}-{} PR Info".format(period[0],period[1])
    wb.create_sheet("{}-{} PR statistics".format(period[0],period[1]),1)
    wb.save(excelfile)
    wb = load_workbook(excelfile)
    ws = wb.worksheets[0]
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 30
    titlefont = Font(name='Arial', size=11, bold=True, italic=False, color='000000')
    titlefill = PatternFill(fill_type="solid",fgColor="FFCC00")
    urlfont = Font(name='Arial', size=11, bold=False, italic=False, underline='single', color='3366FF')
    ws.append(report_header[0])
    for i,j in enumerate(prdata):
        for m,n in enumerate(j):
            if m != 2:
                ws.cell(column=m+1,row=i+2,value=n)
            else:
                ws.cell(column=m+1,row=i+2).hyperlink = n
    for k in range(1,len(report_header[0])+1):
        ws.cell(row=1,column=k).fill = titlefill
        ws.cell(row=1,column=k).font = titlefont
        ws.cell(row=1,column=k).alignment = Alignment(horizontal='center', vertical='center')
    for r in range(2,len(prdata)+2):
        ws.cell(row=r,column=3).font = urlfont
    max_row = ws.max_row
    max_column = ws.max_column
    side = Side(border_style='thin', color='FF000000')
    border = Border(left=side,right=side,top=side,bottom=side)
    for p in range(1, max_row+1):
        for q in range(1, max_column+1):
            ws.cell(p, q).border = border
    ws = wb.worksheets[1]
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.append(report_header[1])
    for row in stat_list:
        ws.append(row)
    for k in range(1,len(report_header[1])+1):
        ws.cell(row=1,column=k).fill = titlefill
        ws.cell(row=1,column=k).font = titlefont
        ws.cell(row=1,column=k).alignment = Alignment(horizontal='center', vertical='center')
    max_row = ws.max_row
    max_column = ws.max_column
    side = Side(border_style='thin', color='FF000000')
    border = Border(left=side,right=side,top=side,bottom=side)
    for p in range(1, max_row+1):
        for q in range(1, max_column+1):
            ws.cell(p, q).border = border
    wb.save(excelfile)



if __name__=="__main__":
    token = para.token
    headers = para.headers
    owner = para.owner
    pkg_list = get_giteepkg(headers,token,owner)
    period = para.period
    prinfo_list = get_PRInfo(pkg_list,headers,token,owner,period)
    report_header = para.report_header
    excelfile = para.excelfile
    create_excelfile(prinfo_list,report_header,excelfile,period)