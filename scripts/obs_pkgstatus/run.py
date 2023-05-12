import requests
import os
import para
from requests.auth import HTTPBasicAuth
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Side, Border
import time

def get_obsdata(url,account,repolist,project):
    pkgs_url = '{}/source/{}'.format(url,project)
    pkgs_resp = requests.get(pkgs_url,auth=HTTPBasicAuth(account['user'],account['password']))
    pkgs_data = pkgs_resp.text
    pkgslist = re.findall('".*"', pkgs_data)
    del pkgslist[0]
    pkgslist = [x.replace('"','') for x in pkgslist]
    print ('pkgs length', len(pkgslist))
    pkginfo_list = []
    for pkg in pkgslist:
        print ('obs package', pkg, pkgslist.index(pkg)+1)
        service_url = '{}/source/{}/{}/_service'.format(url,project,pkg)
        service_resp = requests.get(service_url,auth=HTTPBasicAuth(account['user'],account['password']))
        service_data = service_resp.text
        # print ('service_data', service_data)
        revision_pattern = '<param name="revision">.*</param>'
        try:
            revision = re.search(revision_pattern, service_data).group()
            # print (revision)
            revision = re.search('>.*<', revision).group()[1:-1]
            if len(revision) == 0:
                revision = 'None'
            # print ('git', gitinfo)
            # print ('revision', revision)
        except:
            revision = 'None'
            print ('Cannot get revision!')
        statuslist = []
        for repo in repolist:
            status_url = '{}/build/{}/{}/riscv64/{}/_status'.format(url,project,repo,pkg)
            status_resp = requests.get(status_url,auth=HTTPBasicAuth(account['user'],account['password']))
            status_data = status_resp.text
            # print ('status_data', status_data)
            status = re.search('code=".*"', status_data).group()
            status = status.split('"')[1]
            # print ('status', status)
            statuslist.append(status)
        pkglist = [pkg,revision] + statuslist
        # print ('pkglist>>>', pkglist)
        pkginfo_list.append(pkglist)
        # print ('pkginfo_list>>>', pkginfo_list)
        time.sleep(1)
    return pkginfo_list


def create_excelfile(reportdata,header,excelfile):
    wb = Workbook()
    ws = wb.active
    ws.title = "OBS Packages Info"
    wb.save(excelfile)
    wb = load_workbook(excelfile)
    ws = wb.worksheets[0]
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 60
    for i in range(67, 67+len(repolist)):
        ws.column_dimensions[chr(i)].width = 40
    ws.append(header)
    for row in reportdata:
        ws.append(row)
    max_row = ws.max_row
    max_column = ws.max_column
    side = Side(border_style='thin', color='FF000000')
    border = Border(left=side,right=side,top=side,bottom=side)
    for i in range(1, max_row+1):
        for j in range(1, max_column+1):
            ws.cell(i, j).border = border
    
    wb.save(excelfile)


if __name__=="__main__":
    obs_account = para.obs_account
    obs_url = para.obs_url
    repolist = para.repolist
    obs_project = para.obs_project
    report_data = get_obsdata(obs_url, obs_account, repolist,obs_project)
    excelheader = para.excelheader
    excelfile = para.excelfile
    create_excelfile(report_data, excelheader, excelfile)   